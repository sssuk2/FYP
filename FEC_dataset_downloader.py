from PIL import Image 
import numpy as np
# import pandas as pd
import os
# import tensorflow as tf
# from tensorflow import keras
# from keras.preprocessing.image import ImageDataGenerator, load_img

# from keras.layers import Conv2D, Dense, BatchNormalization, Activation, Dropout, MaxPooling2D, Flatten
# from keras.optimizers import Adam, RMSprop, SGD
# from keras import regularizers
# from keras.callbacks import ModelCheckpoint, CSVLogger, TensorBoard, EarlyStopping, ReduceLROnPlateau
import datetime
import matplotlib 
import matplotlib.pyplot as plt
# from keras.utils import plot_model

#___________
import matplotlib.image as mpimg
import pickle

import requests

#___For FEC dataset
import openpyxl
from pathlib import Path
import glob
import cv2

import urllib
import requests
import shutil

import imageio
from io import BytesIO
import sys

# from skimage import io

# Reading Test File using openpyxl
Tutorial: https://www.tutorialspoint.com/read-and-write-to-an-excel-file-using-python-openpyxl-module

#29 April

#Function to find images format


def ty_len(array):
  print(type(array),len(array),np.shape(array))

def plt_im(image):
  for i in image:
    plt.figure()
    plt.imshow(i)

def p(input):
  print(input)


def find_ext(name):
  #....fucntion to find the extention of the image from the link
  ext = 0
  value = os.path.basename(name)
  # -1 means particular was found in the string
  if value.find('jpg') !=-1: 
    ext = 'jpg'
  elif value.find('JPG') !=-1: 
      ext = 'jpg'
  elif value.find('png') !=-1:
      ext = 'png'
  elif value.find('gif') !=-1:
    ext = 'jpg'
  elif (value.find('jpeg') !=-1 or value.find('JPEG') !=-1 or value.find('Jpeg') !=-1 ):
      ext = 'jpeg'
  else:
    ext = 'jpg'
  return ext 



# Reading excel file using openpyxl
# Column names and index in file 
# 2 : Top-left Column
# 3 : Bottom-Right Clm
# 4 : Top left Row
# 5 : Bottom Right row
# 6 : Image 2 link
# 7 : Top-left Column
# 8 : Bottom-Right Clm
# 9 : Top left Row
# 10 : Bottom Right row
# 11 : Image 3 link
# 12 : Top-left Column
# 13 : Bottom-Right Clm
# 14 : Top left Row
# 15 : Bottom Right row
# 16 : Triplet Type
# 17 : Similarity
# 18 : Six same labels
path = '/content/gdrive/My Drive/Deep Learning FYP/Code/'

# To open the workbook 
# workbook object is created
wb_obj = openpyxl.load_workbook('Train Set.xlsx')
sheet_obj = wb_obj.active

# ptint total number of row & column 
print(sheet_obj.max_row, sheet_obj.max_column)

root_folder = '/content/gdrive/My Drive/Deep Learning FYP/Code/test/'

one_c = 'ONE_CLASS_TRIPLET'
two_c = 'TWO_CLASS_TRIPLET/'
three_c = 'THREE_CLASS_TRIPLET/'

p(os.path.join(root_folder,one_c))



#Directories 
root_folder = '/content/gdrive/My Drive/Deep Learning FYP/Code/train'
# root_folder = '/content/gdrive/My Drive/Deep Learning FYP/Code/test/'

one_c = 'ONE_CLASS_TRIPLET/'
two_c = 'TWO_CLASS_TRIPLET/'
three_c = 'THREE_CLASS_TRIPLET/'

#_____cropped images from bounding boxes ,names and their similarity labels__________

triplet= []
similarity_col = 17
similarity_train = []
r_max = sheet_obj.max_row
triplet_col = 16
triplet_type = []

#no of valid rows and their indecies to keep record in case when file crashes during images downloading   
row_success_count= 0
valid_row_index  = [] 

#no of invalid rows and their indecies 
row_discarded_count = 0
row_discarded_index = [] 


for row in range(2029,8000+1):#in rows
  imgraw= []
  cropped = []
  f_ext = []
  if row%100==0: print(row) #print row at every incrment of 100
  skiprow = False # To check when to skip row
  imglinkcol = [1,6,11] #three images Im1,Im2 and Im3 respectively on col 1, col6 and col 11 respectivley

  #loop through all three images and check if the link is valid
  for linkid in imglinkcol:
    imglink = sheet_obj.cell(row = row, column = linkid).value #extracting image link
    try:
      r = requests.get(imglink, stream = True)
      if r.status_code != 200 or r.status_code ==404 or r.status_code == 410:          #if not valid 
        skiprow = True                  # tells the outer loop to continue
        print('Status code not 200, is', r.status_code)
        row_discarded_count += 1        # counting how many rows will be discarded due to invalid image links
        row_discarded_index.append(row) #index of row in case to double check the erros 
        break                           # dont process subsequent images in the triplet
      else:
        img = Image.open(BytesIO(r.content))
        img = np.array(img)
        # checking the number of channels in the images as plt.imsave can't save images of channels less then 3
        # if ((img.shape[2] != 3) and (img.shape[2] !=4)):
        #   skiprow = True
        #   break
        # else:
        imgraw.append(img)        
    except Exception as e:              # Raise exception if there is any certificate error as image couldn't be retrieved
      print('Failed to download image: ', e)
      row_discarded_count += 1          # counting how many rows will be discarded due to invalid image links
      row_discarded_index.append(row)   #index of row in case to double check the errors 
      skiprow = True
      break
    
  if skiprow: continue                  # skip row if any image cannot be downloaded   
  valid_row_index.append(row)  
  row_success_count += 1  
  # recording similarity between the two images
  sim = (int)(sheet_obj.cell(row = row, column = similarity_col).value)
  similarity_train.append(sim)

  # stare processing
  #cropping images
  for i,linkid in enumerate(imglinkcol):
    
    if linkid==1:   #1st Image
      h,w = np.shape(imgraw[i])[:2] #accessing height and width of image 1 
      # l & r is left and right fraction of widht to crop the image respectively
      left   =  (int)((sheet_obj.cell(row = row, column =  linkid + 1  ).value)  * w)
      right   = (int)((sheet_obj.cell(row = row, column = linkid + 2  ).value)  * w)
      # t & b is top and bottom fraction of height to crop the image respectively
      top   =   (int)((sheet_obj.cell(row = row, column = linkid + 3  ).value)  * h)
      bottom   = (int)((sheet_obj.cell(row = row, column = linkid + 4  ).value)  * h)
      # crop images in imgraw
        # bounding box is comprised (top, bottom ,left ,right) cordinates
        # crop images will be fraction of width and height of oringal image
        # where new height = value of pixel from top to bottom
        # and   new width =  value of pixels from left to right
      cropped.append(imgraw[i][top:bottom,left:right])
      f_ext.append(find_ext(sheet_obj.cell(row = row, column = linkid).value))

    elif linkid==6:     # 2nd image
      h,w = np.shape(imgraw[i])[:2] #accessing height and width of image 1 
      # l & r is left and right fraction of widht to crop the image respectively
      left   =  (int)((sheet_obj.cell(row = row, column =  linkid + 1  ).value)  * w)
      right   = (int)((sheet_obj.cell(row = row, column = linkid + 2  ).value)  * w)
      # t & b is top and bottom fraction of height to crop the image respectively
      top   =   (int)((sheet_obj.cell(row = row, column = linkid + 3  ).value)  * h)
      bottom   = (int)((sheet_obj.cell(row = row, column = linkid + 4  ).value)  * h)
      cropped.append(imgraw[i][top:bottom,left:right])
      f_ext.append(find_ext(sheet_obj.cell(row = row, column = linkid).value))

    elif linkid==11:  #3rd Image 
      h,w = np.shape(imgraw[i])[:2] #accessing height and width of image 1 
      # l & r is left and right fraction of widht to crop the image respectively
      left   =  (int)((sheet_obj.cell(row = row, column =  linkid + 1  ).value)  * w)
      right   = (int)((sheet_obj.cell(row = row, column = linkid + 2  ).value)  * w)
      # t & b is top and bottom fraction of height to crop the image respectively
      top   =   (int)((sheet_obj.cell(row = row, column = linkid + 3  ).value)  * h)
      bottom   = (int)((sheet_obj.cell(row = row, column = linkid + 4  ).value)  * h)
      cropped.append(imgraw[i][top:bottom,left:right])
      f_ext.append(find_ext(sheet_obj.cell(row = row, column = linkid).value))

  # Check Triplet type and save images in respective folders 
  os.chdir(root_folder)
  if (sheet_obj.cell(row = row, column = triplet_col).value == 'ONE_CLASS_TRIPLET'):
    #join path to one class trip folder
    os.chdir(os.path.join(root_folder,one_c))

    #make new folder with row index and join the path with triplet folder
    n_row =  'Row '+ str(row)
    os.makedirs(n_row)  
    os.chdir(os.path.join(root_folder,one_c,n_row))

    if sim == 1:
      # image naming convention
        # NAME : '1x.jpg' 
          # first digit is just image number
          # Second digit Important as its the similiarity score
            # x represent the image that isn't similar to other two imgages 
      plt.imsave('1x',cropped[0], format = f_ext[0])
      plt.imsave('21',cropped[1], format = f_ext[1])
      plt.imsave('31',cropped[2], format = f_ext[2])
    elif sim ==2:
      plt.imsave('12',cropped[0], format =f_ext[0])
      plt.imsave('2x',cropped[1], format =f_ext[1])
      plt.imsave('32',cropped[2], format =f_ext[2])
    elif sim ==3:
      plt.imsave('13',cropped[0], format =f_ext[0])
      plt.imsave('23',cropped[1], format =f_ext[1])
      plt.imsave('2x',cropped[2], format =f_ext[2])
  elif (sheet_obj.cell(row = row, column = triplet_col).value == 'TWO_CLASS_TRIPLET'):
    os.chdir(os.path.join(root_folder,two_c))

    #make new folder with row index and join the path with triplet folder
    n_row =  'Row '+ str(row)
    os.makedirs(n_row)
    os.chdir(os.path.join(root_folder,two_c,n_row))
    if sim == 1:
      plt.imsave('1x',cropped[0], format =f_ext[0])
      plt.imsave('21',cropped[1], format =f_ext[1])
      plt.imsave('31',cropped[2], format =f_ext[2])
    elif sim ==2:
      plt.imsave('12',cropped[0], format =f_ext[0])
      plt.imsave('2x',cropped[1], format =f_ext[1])
      plt.imsave('32',cropped[2], format =f_ext[2])
    elif sim ==3:
      plt.imsave('13',cropped[0], format =f_ext[0])
      plt.imsave('23',cropped[1], format =f_ext[1])
      plt.imsave('3x',cropped[2], format =f_ext[2])
  elif (sheet_obj.cell(row = row, column = triplet_col).value == 'THREE_CLASS_TRIPLET'):
    #join path to one class trip folder
    os.chdir(os.path.join(root_folder,three_c))
    
    #make new folder with row index and join the path with triplet folder
    n_row =  'Row '+ str(row)
    os.makedirs(n_row)
    os.chdir(os.path.join(root_folder,three_c,n_row))
    # os.path.join(root_folder,one_c,n_row)
    if sim == 1:
      plt.imsave('1x',cropped[0], format = f_ext[0])
      plt.imsave('21',cropped[1], format = f_ext[1])
      plt.imsave('31',cropped[2], format = f_ext[2])
    elif sim ==2:
      plt.imsave('12',cropped[0], format = f_ext[0])
      plt.imsave('2x',cropped[1], format = f_ext[1])
      plt.imsave('32',cropped[2], format = f_ext[2])
    elif sim ==3:
      plt.imsave('13',cropped[0], format = f_ext[0])
      plt.imsave('23',cropped[1], format = f_ext[1])
      plt.imsave('3x',cropped[2], format = f_ext[2])
     
          
   
   
print('done with the loop') 
print('invlaide rows',row_discarded_count)
print('valid rows', row_success_count)
print('len of imgraw is',len(imgraw))
print('len of cropped is', len(cropped))
print('len of similarity scores', len(similarity_train))
print('last index where loop stoped',valid_row_index[-1])

